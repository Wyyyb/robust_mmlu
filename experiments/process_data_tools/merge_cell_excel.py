import pandas as pd
from openpyxl import load_workbook


def merge_excel_columns(excel_path):
    # 读取Excel文件到pandas DataFrame
    df = pd.read_excel(excel_path, engine='openpyxl')

    # 加载工作簿用于合并单元格
    book = load_workbook(excel_path)
    sheet = book.active

    # 对于每一列进行合并操作
    for col_idx in range(1, df.shape[1] + 1):
        if col_idx not in [3, 4]:
            continue
        previous_value = None
        merge_start = None
        for row_idx in range(2, df.shape[0] + 2):  # 从第二行开始（排除表头）
            # 读取单元格的值
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            # 如果单元格的值与前一个值相同，则可能需要合并
            if cell_value == previous_value:
                if merge_start is None:
                    merge_start = row_idx - 1  # 记录合并的起始行
            else:
                # 如果值不同并且之前有记录的合并起始行，则执行合并操作
                if merge_start is not None and merge_start < row_idx - 1:
                    # 合并单元格
                    sheet.merge_cells(start_row=merge_start, start_column=col_idx, end_row=row_idx - 1,
                                      end_column=col_idx)
                merge_start = row_idx  # 重置起始行为当前行
            previous_value = cell_value
        # 检查是否需要在列的末尾进行合并
        if merge_start is not None and merge_start < df.shape[0] + 1:
            sheet.merge_cells(start_row=merge_start, start_column=col_idx, end_row=df.shape[0] + 1, end_column=col_idx)

    # 保存更改后的Excel文件
    book.save(excel_path)

# 使用函数示例
merge_excel_columns('../data/mmlu_sta.xltx')


